from __future__ import annotations

import json
import re
import sys
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

import fitz
import openpyxl

try:
    from rhwp import parse as parse_hwp
except Exception:  # pragma: no cover
    parse_hwp = None


ROOT = Path(r"C:\Users\chris\Desktop\suprema-platform")
INPUT_XLSX = ROOT / "outputs" / "2028_입시계획" / "2028_최종결과.xlsx"
OUTPUT_XLSX = ROOT / "outputs" / "2028_numeric_18_repaired_v3.xlsx"
PDF_ROOT = ROOT / "data" / "2028_pdf_extracted"
RAW_ROOT = ROOT / "data" / "pdf_raw_text"

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="ignore")


def normalize(text: object) -> str:
    value = str(text or "")
    value = value.replace("\u0000", "")
    value = re.sub(r"\([^)]*\)", "", value)
    value = re.sub(r"\[[^\]]*\]", "", value)
    value = value.replace(" ", "")
    value = value.replace("|", "")
    value = value.replace("\t", "")
    value = value.replace("·", "")
    value = value.replace("-", "")
    value = value.replace("'", "")
    value = value.replace("“", "").replace("”", "")
    value = value.replace("(", "").replace(")", "")
    return value


def university_aliases(university: str) -> list[str]:
    base = re.sub(r"\([^)]*\)", "", university)
    aliases: list[str] = []
    for candidate in (university, base):
        candidate = candidate.strip()
        if not candidate:
            continue
        normalized = normalize(candidate)
        if normalized and normalized not in aliases:
            aliases.append(normalized)
        if candidate.endswith("대학교"):
            short = candidate[:-3] + "대"
            normalized_short = normalize(short)
            if normalized_short and normalized_short not in aliases:
                aliases.append(normalized_short)
        if candidate.endswith("대학"):
            stripped = candidate[:-2]
            normalized_short = normalize(stripped)
            if normalized_short and normalized_short not in aliases:
                aliases.append(normalized_short)
    return aliases


def numeric_values(text: str) -> list[int]:
    values: list[int] = []
    for token in re.findall(r"\d{1,3}(?:,\d{3})*|\d+", text):
        try:
            values.append(int(token.replace(",", "")))
        except ValueError:
            continue
    return values


def row_kind(name: str, admission_type: str) -> str:
    text = f"{name or ''} {admission_type or ''}"
    if any(key in text for key in ("정시", "수능", "수능위주", "교과이수", "실기")):
        return "jungsi"
    if any(key in text for key in ("논술", "학생부", "면접", "종합", "기회균형", "특별전형", "추천")):
        return "susi"
    return "susi"


def type_position(name: str, admission_type: str, count_len: int) -> int | None:
    text = f"{name or ''} {admission_type or ''}"
    if count_len <= 0:
        return None
    if "기회균형" in text:
        return 2 if count_len >= 3 else count_len - 1
    if "추천" in text:
        return min(1, count_len - 1)
    if "정시" in text or "수능" in text:
        return count_len - 1
    if "교과" in text:
        return 0
    if "종합" in text:
        return 0 if count_len == 1 else min(1, count_len - 1)
    return 0


def expand_table_rows(rows: list[list[object]]) -> list[str]:
    lines: list[str] = []
    for row in rows:
        cells = ["" if cell is None else str(cell).strip() for cell in row]
        if not any(cells):
            continue
        split_cells = [cell.splitlines() if cell else [""] for cell in cells]
        height = max(len(parts) for parts in split_cells)
        for idx in range(height):
            parts = [parts[idx].strip() if idx < len(parts) else "" for parts in split_cells]
            line = "\t".join(parts).strip()
            if line:
                lines.append(line)
    return lines


def page_texts_from_pdf(path: Path) -> list[tuple[int, str]]:
    doc = fitz.open(path)
    pages: list[tuple[int, str]] = []
    for index, page in enumerate(doc, start=1):
        lines = []
        table_lines = []
        try:
            finder = page.find_tables()
            for table in finder.tables[:4]:
                try:
                    table_lines.extend(expand_table_rows(table.extract()))
                except Exception:
                    continue
        except Exception:
            pass
        if table_lines:
            lines.extend(table_lines)
        text = page.get_text("text") or ""
        if text.strip():
            lines.extend(line.strip() for line in text.splitlines() if line.strip())
        if not lines:
            blocks = page.get_text("blocks") or ""
            if isinstance(blocks, str) and blocks.strip():
                lines.append(blocks.strip())
            else:
                lines.append(str(blocks))
        pages.append((index, "\n".join(line for line in lines if line)))
    return pages


def page_texts_from_json(path: Path) -> list[tuple[int, str]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return [(int(page["page"]), str(page.get("text") or "")) for page in data.get("pages", [])]


def page_texts_from_xlsx(path: Path) -> list[tuple[int, str]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active
    lines = []
    for row in sheet.iter_rows(values_only=True):
        values = [str(value).strip() for value in row if value not in (None, "")]
        if values:
            lines.append(" ".join(values))
    return [(1, "\n".join(lines))]


def page_texts_from_hwp(path: Path) -> list[tuple[int, str]]:
    if parse_hwp is None:
        return []
    try:
        parsed = parse_hwp(str(path))
        text = parsed.extract_text() or ""
        return [(1, text)]
    except Exception:
        return []


def source_pages(path: Path) -> list[tuple[int, str]]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return page_texts_from_json(path)
    if suffix == ".pdf":
        return page_texts_from_pdf(path)
    if suffix == ".xlsx":
        return page_texts_from_xlsx(path)
    if suffix == ".hwp":
        return page_texts_from_hwp(path)
    return []


def candidate_sources(university: str) -> list[Path]:
    aliases = university_aliases(university)
    scored: list[tuple[float, Path]] = []
    all_files = [p for p in PDF_ROOT.rglob("*") if p.is_file()]
    raw_files = [p for p in RAW_ROOT.glob("*.json")]

    def score_path(path: Path) -> float:
        stem = normalize(path.stem)
        score = 0.0
        for alias in aliases:
            if not alias:
                continue
            if stem == alias:
                score = max(score, 100.0 + len(alias) / 100.0)
            elif stem.startswith(alias):
                score = max(score, 60.0 + len(alias) / 100.0)
            elif alias in stem:
                score = max(score, 30.0 + len(alias) / 100.0)
            else:
                score = max(score, SequenceMatcher(None, alias, stem).ratio() * 10.0)
        return score

    for path in all_files:
        score = score_path(path)
        if score > 0:
            scored.append((score, path))

    # Prefer raw text exports if they clearly match.
    for path in raw_files:
        stem = normalize(path.stem)
        for alias in aliases:
            if alias and (stem == alias or stem.startswith(alias) or alias in stem):
                scored.append((150.0 + len(alias) / 100.0, path))
                break

    scored.sort(key=lambda item: (item[0], -len(item[1].name)), reverse=True)
    result: list[Path] = []
    seen: set[Path] = set()
    for _, path in scored:
        if path not in seen:
            seen.add(path)
            result.append(path)
    return result


def section_kind(text: str) -> str:
    if "정시" in text or "수능" in text:
        return "jungsi"
    if "정원외" in text or "정원 외" in text:
        return "outside"
    return "susi"


def section_match_score(page_text: str, target_kind: str) -> int:
    kind = section_kind(page_text)
    if kind == target_kind:
        return 3
    if target_kind == "outside" and kind == "susi":
        return 1
    return 0


def extract_count_for_row(university_pages: list[tuple[int, str]], dept: str, admission_type: str, name: str) -> tuple[int | None, str | None]:
    dept_key = normalize(dept)
    target_kind = row_kind(name, admission_type)
    best: tuple[tuple[int, int, int, int], int | None, str | None] | None = None

    for page_no, text in university_pages:
        lines = [re.sub(r"\s+", " ", line).strip() for line in str(text).splitlines() if line.strip()]
        if not lines:
            continue
        header_text = " ".join(lines[:8])
        if section_match_score(header_text, target_kind) == 0:
            continue
        for idx, line in enumerate(lines):
            line_key = normalize(line)
            if dept_key not in line_key:
                continue
            window_text = line
            if not numeric_values(window_text):
                for offset in range(1, 6):
                    if idx + offset >= len(lines):
                        break
                    next_line = lines[idx + offset]
                    if any(token in next_line for token in ("학생부", "정시", "수시", "논술", "실기", "지원자격", "전형방법")):
                        break
                    window_text += " " + next_line
            nums = numeric_values(window_text)
            if not nums:
                continue
            if "\t" not in line and "|" not in line:
                sentence_like = any(
                    token in window_text
                    for token in (
                        "면접",
                        "지원자격",
                        "전형방법",
                        "수능최저",
                        "학력",
                        "평가",
                        "서류",
                        "진행",
                        "이수",
                        "기준",
                        "안내",
                        "과목",
                        "등급",
                        "이내",
                        "선발",
                        "합격",
                        "추천",
                        "포함",
                        "최저",
                        "교과과정",
                        "교육과정",
                        "학기",
                        "단위",
                    )
                )
                if sentence_like and len(nums) == 1:
                    continue
                if len(window_text) > 90 and len(nums) <= 1:
                    continue
                if any(token in window_text for token in ("개정", "학년도", "연도", "년도")) and any(
                    1900 <= num <= 2035 for num in nums
                ):
                    nums = [num for num in nums if num < 1900 or num > 2035]
                    if not nums:
                        continue
            position = type_position(name, admission_type, len(nums))
            if position is None:
                continue
            if position >= len(nums):
                position = len(nums) - 1
            value = nums[position]
            score = (
                section_match_score(header_text, target_kind),
                1 if name and normalize(name) in line_key else 0,
                -len(nums),
                -len(line_key),
            )
            candidate = (score, value, line)
            if best is None or candidate[0] > best[0]:
                best = candidate

    if best is None:
        return None, None
    return best[1], best[2]


def main() -> None:
    workbook = openpyxl.load_workbook(INPUT_XLSX)
    sheet = workbook.active

    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 9).value = None

    rows_by_university: dict[str, list[int]] = {}
    for row in range(2, sheet.max_row + 1):
        university = str(sheet.cell(row, 3).value or "").strip()
        if university:
            rows_by_university.setdefault(university, []).append(row)

    source_cache: dict[Path, list[tuple[int, str]]] = {}
    row_updates: dict[int, int] = {}
    debug_rows = []

    for university, row_indexes in rows_by_university.items():
        sources = candidate_sources(university)
        if not sources:
            continue

        pages: list[tuple[int, str]] = []
        chosen_source: Path | None = None
        for source in sources[:4]:
            if source not in source_cache:
                try:
                    source_cache[source] = source_pages(source)
                except Exception:
                    source_cache[source] = []
            pages = source_cache[source]
            if pages:
                chosen_source = source
                break
        if not pages or chosen_source is None:
            continue

        for row in row_indexes:
            dept = str(sheet.cell(row, 5).value or "").strip()
            admission_type = str(sheet.cell(row, 6).value or "").strip()
            name = str(sheet.cell(row, 7).value or "").strip()
            count, matched_line = extract_count_for_row(pages, dept, admission_type, name)
            if count is not None:
                row_updates[row] = count
                if len(debug_rows) < 30:
                    debug_rows.append([row, university, dept, admission_type, name, count, matched_line, str(chosen_source)])

    for row, value in row_updates.items():
        sheet.cell(row, 9).value = value

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_XLSX)

    report = {
        "input": str(INPUT_XLSX),
        "output": str(OUTPUT_XLSX),
        "updated_rows": len(row_updates),
        "debug": debug_rows,
    }
    (ROOT / "_tmp").mkdir(exist_ok=True)
    (ROOT / "_tmp" / "repair_moc_column_v2_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
