from __future__ import annotations

import json
import re
import subprocess
import sys
from pathlib import Path

import fitz
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font


ROOT = Path(r"C:\Users\chris\Desktop\suprema-platform")
TESSERACT = Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe")
TMP_DIR = ROOT / ".tmp" / "adiga_ocr_pages_extract"

KEYWORDS = ("모집인원", "모집단위", "입학정원", "전형명")


def safe_text(value):
    if not isinstance(value, str):
        return value
    value = ILLEGAL_CHARACTERS_RE.sub(" ", value)
    value = "".join(
        ch
        for ch in value
        if ch in "\t\n\r"
        or 0x20 <= ord(ch) <= 0xD7FF
        or 0xE000 <= ord(ch) <= 0xFFFD
        or 0x10000 <= ord(ch) <= 0x10FFFF
    )
    return re.sub(r"\s+", " ", value).strip()


def ocr_page(page: fitz.Page, pdf_stem: str, page_index: int) -> str:
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    image_path = TMP_DIR / f"{pdf_stem}_p{page_index + 1}.png"
    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
    pix.save(image_path)
    try:
        result = subprocess.run(
            [str(TESSERACT), str(image_path), "stdout", "-l", "kor+eng"],
            capture_output=True,
            text=False,
            check=False,
        )
        return (result.stdout or b"").decode("utf-8", errors="ignore")
    finally:
        if image_path.exists():
            image_path.unlink()


def extract_candidates(text: str) -> list[str]:
    text = re.sub(r"\s+", " ", text or "")
    candidates: list[str] = []
    patterns = [
        r"모집인원[^0-9]{0,20}([0-9][0-9,]*)",
        r"모집인원[^0-9]{0,40}([0-9][0-9,]*)",
        r"모집인원\s*[:：]?\s*([0-9][0-9,]*)",
        r"입학정원[^0-9]{0,20}([0-9][0-9,]*)",
        r"모집단위[^0-9]{0,20}([0-9][0-9,]*)",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, text):
            value = match.group(1).replace(",", "")
            if not value.isdigit():
                continue
            number = int(value)
            if 1900 <= number <= 2100:
                continue
            candidates.append(str(number))
        if candidates:
            break
    seen = set()
    unique: list[str] = []
    for value in candidates:
        if value not in seen:
            seen.add(value)
            unique.append(value)
    return unique


def choose_value(candidates: list[str]) -> str:
    return candidates[0] if len(candidates) == 1 else ""


def main() -> None:
    if len(sys.argv) < 3:
        raise SystemExit("usage: extract_candidates_from_summary.py <summary.json> <output.xlsx>")

    summary_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    summary = json.loads(summary_path.read_text(encoding="utf-8"))

    workbook = Workbook()
    ws_summary = workbook.active
    ws_summary.title = "summary"
    ws_pages = workbook.create_sheet("pages")
    ws_review = workbook.create_sheet("review")

    summary_headers = ["school", "source_files", "bundle_pages", "bundle_path", "hit_pages_total", "pages_with_moc", "first_value", "review_pages"]
    page_headers = ["school", "member", "page", "keywords", "candidates", "selected", "source_kind", "snippet"]
    review_headers = ["school", "member", "page", "issue", "candidates", "snippet"]

    ws_summary.append(summary_headers)
    ws_pages.append(page_headers)
    ws_review.append(review_headers)
    for ws in (ws_summary, ws_pages, ws_review):
        for cell in ws[1]:
            cell.font = Font(bold=True)

    for entry in summary:
        school = safe_text(entry.get("school", ""))
        bundle_value = entry.get("bundle_path") or ""
        if not bundle_value:
            ws_summary.append([school, entry.get("source_files", 0), entry.get("bundle_pages", 0), "", 0, 0, "", "no_bundle"])
            continue

        bundle_path = Path(bundle_value)
        if not bundle_path.exists():
            ws_summary.append([school, entry.get("source_files", 0), entry.get("bundle_pages", 0), safe_text(bundle_value), 0, 0, "", "missing_bundle"])
            continue

        bundle = fitz.open(bundle_path)
        hit_pages_total = 0
        pages_with_moc = 0
        first_value = ""
        review_pages = 0
        files = entry.get("files") or []
        first_member = safe_text(files[0].get("member", "")) if files else ""

        for page_index in range(bundle.page_count):
            page = bundle[page_index]
            text = page.get_text("text")
            source_kind = "text"
            if not text.strip():
                text = ocr_page(page, bundle_path.stem, page_index)
                source_kind = "ocr"
            text = safe_text(text)
            keywords = [keyword for keyword in KEYWORDS if keyword in text]
            if not keywords:
                continue

            hit_pages_total += 1
            idx = min((text.find(keyword) for keyword in keywords if text.find(keyword) >= 0), default=0)
            snippet = text[max(0, idx - 160): idx + 240]
            candidates = extract_candidates(text if "모집인원" in keywords else snippet)
            selected = choose_value(candidates)
            if "모집인원" in keywords:
                pages_with_moc += 1
                if not first_value and selected:
                    first_value = selected
                if len(candidates) != 1:
                    review_pages += 1
                    ws_review.append([school, first_member, page_index + 1, "ambiguous" if candidates else "no_candidate", ", ".join(candidates), snippet])

            ws_pages.append([school, first_member, page_index + 1, ", ".join(keywords), ", ".join(candidates), selected, source_kind, snippet])

        ws_summary.append([school, entry.get("source_files", 0), entry.get("bundle_pages", 0), safe_text(bundle_value), hit_pages_total, pages_with_moc, first_value, review_pages])

    for ws in (ws_summary, ws_pages, ws_review):
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    widths = {
        ws_summary: [24, 12, 12, 72, 14, 14, 12, 14],
        ws_pages: [20, 36, 8, 18, 18, 12, 10, 100],
        ws_review: [20, 36, 8, 16, 18, 100],
    }
    for ws, width_list in widths.items():
        for idx, width in enumerate(width_list, start=1):
            ws.column_dimensions[chr(64 + idx)].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(json.dumps({"output": str(output_path), "summary_rows": ws_summary.max_row - 1, "page_rows": ws_pages.max_row - 1, "review_rows": ws_review.max_row - 1}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
