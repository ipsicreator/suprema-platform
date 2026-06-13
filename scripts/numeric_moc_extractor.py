from __future__ import annotations

import json
import re
import os
import shutil
import subprocess
import tempfile
import sys
from collections import Counter, defaultdict
from difflib import SequenceMatcher
from pathlib import Path

import fitz
import openpyxl

ROOT = Path(r"C:\Users\chris\Desktop\suprema-platform")
PYLIBS = ROOT / ".cache" / "pylibs"
if str(PYLIBS) not in sys.path:
    sys.path.insert(0, str(PYLIBS))
WORKBOOK_PATH = Path(r"C:\Users\chris\Desktop\2028학년도_입시계획_최종.xlsx")
OUTPUT_PATH = ROOT / "outputs" / "numeric_moc_output.xlsx"
DESKTOP_OUTPUT_PATH = Path(r"C:\Users\chris\Desktop\2028학년도_입시계획_최종_숫자만.xlsx")
SOURCE_ROOT = ROOT / "data" / "2028_pdf_extracted"
STATUS_PATH = ROOT / "_tmp" / "numeric_moc_status.json"
DONE_PATH = ROOT / "_tmp" / "numeric_moc.done"
TESSERACT = Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe")
TESSDATA_DIR = ROOT / ".cache" / "ocr" / "tessdata"

try:
    from rhwp import parse as parse_hwp
except Exception:
    parse_hwp = None


def normalize_text(value: str) -> str:
    text = str(value or "")
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"\[[^\]]*\]", "", text)
    text = re.sub(r"[\s\-._/|ㅣ,:;\'\"`~!@#$%^&*+=<>?\\{}()]+", "", text)
    return text


def canonical_university_text(value: str) -> str:
    text = normalize_text(value)
    text = text.replace("국립", "")
    text = text.replace("캠퍼스", "")
    text = re.sub(r"(대학교|대학)$", "", text)
    text = re.sub(r"대$", "", text)
    return text


def parse_number(text: str) -> int | None:
    if not text:
        return None
    candidates = re.findall(r"\d{1,4}(?:,\d{3})*", text.replace(" ", ""))
    for candidate in reversed(candidates):
        value = candidate.replace(",", "")
        if value.isdigit():
            return int(value)
    return None


def plausible_count(value: int | None) -> bool:
    return value is not None and 0 < value <= 1000


def source_key(path: Path) -> str:
    stem = path.stem
    prefix = stem
    prefix = re.split(r"2028|2027|2026", prefix, maxsplit=1)[0]
    prefix = prefix.replace("?", "")
    prefix = prefix.replace("|", "")
    prefix = prefix.replace("ㅣ", "")
    prefix = prefix.replace("l", "")
    return canonical_university_text(prefix)


def university_key(value: str) -> str:
    return canonical_university_text(value)


def university_aliases(value: str) -> set[str]:
    base = university_key(value)
    aliases = {base}
    replacements = [
        ("대학교", "대"),
        ("대학교", ""),
        ("대학", "대"),
        ("(미래)", "미래"),
        ("(미래)", "미래캠퍼스"),
        ("(춘천)", ""),
        ("(춘천)", "춘천캠퍼스"),
        ("(도계)", ""),
        ("(도계)", "도계캠퍼스"),
        ("(세종)", ""),
        ("(세종)", "세종캠퍼스"),
        ("(WISE)", "WISE"),
        ("(ERICA)", "ERICA"),
    ]
    for old, new in replacements:
        aliases.add(base.replace(old, new))
    if base.endswith("대학교"):
        stem = base[:-4]
        aliases.add(stem + "대")
        aliases.add(stem)
    manual_aliases = {
        "충북대학교": {"충북대"},
        "전남대학교": {"전남대"},
        "한경대학교": {"한경대", "한경국립대"},
        "협성대학교": {"협성대"},
        "목포해양대학교": {"목포해양대"},
        "중부대학교": {"중부대"},
        "한라대학교": {"한라대"},
        "을지대학교": {"을지대"},
        "중원대학교": {"중원대"},
        "한국기술교육대학교": {"한국기술교육대"},
        "금오공과대학교": {"금오공과대"},
        "원광대학교": {"원광대"},
        "가야대학교": {"가야대"},
        "경상대학교": {"경상대", "경상국립대"},
        "공주대학교": {"공주대"},
        "한밭대학교": {"한밭대"},
        "한국해양대학교": {"한국해양대"},
        "서울여자대학교": {"서울여대"},
        "경국대학교": {"국립경국대", "경국대"},
        "연세대학교(미래)": {"연세대미래캠퍼스", "연세대 미래캠퍼스"},
        "강원대학교(춘천)": {"강원대춘천캠퍼스", "강원대 춘천캠퍼스"},
        "강원대학교(도계)": {"강원대도계캠퍼스", "강원대 도계캠퍼스"},
        "고려대학교(세종)": {"고려대세종캠퍼스", "고려대 세종캠퍼스", "고려대세종"},
        "동국대학교(WISE)": {"동국대WISE", "동국대 WISE"},
        "한양대학교(ERICA)": {"한양대ERICA", "한양대 ERICA"},
    }
    aliases.update(manual_aliases.get(base, set()))
    return {alias for alias in aliases if alias}


def department_key(value: str) -> str:
    text = normalize_text(value)
    text = re.sub(r"\d+$", "", text)
    return text


def loose_department_key(value: str) -> str:
    text = department_key(value)
    text = text.replace("캠퍼스", "")
    text = re.sub(r"(학과|학부|전공|계열|부)$", "", text)
    text = re.sub(r"(학부|학과|전공|계열)+", "", text)
    return text


def page_text_pdf(page: fitz.Page, page_index: int, workdir: Path) -> str:
    text = page.get_text("text") or ""
    if len(text.strip()) >= 80:
        return text
    image_path = workdir / f"page_{page_index + 1}.png"
    base = workdir / f"page_{page_index + 1}"
    page.get_pixmap(matrix=fitz.Matrix(2.2, 2.2), alpha=False).save(str(image_path))
    subprocess.run(
        [
            str(TESSERACT),
            str(image_path),
            str(base),
            "--psm",
            "4",
            "-l",
            "kor+eng",
            "--tessdata-dir",
            str(TESSDATA_DIR),
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="ignore",
    )
    txt_path = base.with_suffix(".txt")
    if txt_path.exists():
        return txt_path.read_text(encoding="utf-8", errors="ignore")
    return text


def page_text_hwp(path: Path) -> str:
    if parse_hwp is None:
        return ""
    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir) / "input.hwp"
            shutil.copyfile(path, temp_path)
            doc = parse_hwp(str(temp_path))
            ir = json.loads(doc.to_ir_json())
            parts: list[str] = []
            for node in ir.get("body", []):
                if not isinstance(node, dict):
                    continue
                text = str(node.get("text") or "").strip()
                if text:
                    parts.append(text)
            text = "\n".join(parts)
            return text or ""
    except Exception:
        return ""


def xlsx_target_column(rows: list[tuple[object, ...]]) -> int | None:
    hints = (
        "모집인원",
        "모집 인원",
        "정원내",
        "정원+이월",
        "조정 모집 인원",
        "조정모집인원",
        "모집단위별 모집인원",
    )
    for row in rows[:10]:
        if sum(1 for value in row if value not in (None, "")) < 2:
            continue
        for index, value in enumerate(row, start=1):
            text = normalize_text(value)
            if not text:
                continue
            if any(normalize_text(hint) in text for hint in hints):
                return index
    return None


def xlsx_row_label(row: tuple[object, ...], target_column: int | None) -> str | None:
    skip = {
        "대학",
        "캠퍼스",
        "단과대학",
        "단과대학명",
        "모집단위",
        "모집단위명",
        "학부",
        "학과",
        "전공",
        "소계",
        "합계",
        "계",
    }
    candidates: list[tuple[int, str]] = []
    for index, value in enumerate(row[:8], start=1):
        if target_column is not None and index == target_column:
            continue
        text = str(value or "").strip()
        if not text or parse_number(text) is not None:
            continue
        text = re.sub(r"\s+", " ", text)
        normalized = department_key(text)
        if not normalized or normalized in skip:
            continue
        if any(token in normalized for token in ("학과", "학부", "전공", "계열", "부", "원")):
            return normalized
        candidates.append((len(normalized), normalized))
    if candidates:
        candidates.sort(reverse=True)
        return candidates[0][1]
    return None


def pairs_from_xlsx(path: Path) -> list[tuple[str, int]]:
    pairs: list[tuple[str, int]] = []
    try:
        workbook = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return pairs
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        target_column = xlsx_target_column(rows)
        if target_column is None:
            continue
        for row in rows:
            label = xlsx_row_label(row, target_column)
            if not label:
                continue
            value = row[target_column - 1] if target_column - 1 < len(row) else None
            number = parse_number(str(value or ""))
            if not plausible_count(number):
                continue
            pairs.append((label, number))
    return pairs


def parse_pairs(text: str) -> list[tuple[str, int]]:
    pairs: list[tuple[str, int]] = []
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines() if line.strip()]
    for index, line in enumerate(lines):
        if not re.search(r"[가-힣A-Za-z]", line):
            continue
        direct = re.search(r"^(.+?)\s+(\d{1,4}(?:,\d{3})*)$", line)
        if direct:
            number = parse_number(direct.group(2))
            if plausible_count(number):
                pairs.append((department_key(direct.group(1)), number))
                continue
        window = " ".join(lines[index : index + 5])
        number = parse_number(window)
        if plausible_count(number):
            pairs.append((department_key(line), number))
    return pairs


def build_source_maps(target_university: str = "") -> dict[str, dict[str, int]]:
    source_label_counts: dict[str, dict[str, Counter[int]]] = defaultdict(lambda: defaultdict(Counter))
    target_key = university_key(target_university)
    with tempfile.TemporaryDirectory() as temp_dir:
        workdir = Path(temp_dir)
        for path in sorted(SOURCE_ROOT.rglob("*")):
            if path.suffix.lower() not in {".pdf", ".hwp", ".xlsx"}:
                continue
            key = source_key(path)
            if target_key:
                source_candidate = university_key(key)
                if target_key not in source_candidate and source_candidate not in target_key:
                    if SequenceMatcher(None, target_key, source_candidate).ratio() < 0.6:
                        continue
            try:
                if path.suffix.lower() == ".pdf":
                    doc = fitz.open(str(path))
                    texts = [page_text_pdf(page, page_index, workdir) for page_index, page in enumerate(doc)]
                    text = "\n".join(texts)
                    pairs = parse_pairs(text)
                elif path.suffix.lower() == ".hwp":
                    text = page_text_hwp(path)
                    pairs = parse_pairs(text)
                else:
                    pairs = pairs_from_xlsx(path)
            except Exception:
                continue
            if not pairs:
                continue
            for label, number in pairs:
                if label:
                    source_label_counts[key][label][number] += 1
    source_maps: dict[str, dict[str, int]] = {}
    for key, label_counts in source_label_counts.items():
        source_maps[key] = {
            label: counter.most_common(1)[0][0]
            for label, counter in label_counts.items()
            if counter
        }
    return source_maps


def best_source(university: str, source_maps: dict[str, dict[str, int]]) -> str | None:
    aliases = university_aliases(university)
    if not aliases:
        return None
    for alias in aliases:
        if alias in source_maps:
            return alias
    for candidate in source_maps:
        if any(alias in candidate or candidate in alias for alias in aliases):
            return candidate
    best_candidate = None
    best_score = 0.0
    for candidate in source_maps:
        score = max(SequenceMatcher(None, alias, candidate).ratio() for alias in aliases)
        if score > best_score:
            best_score = score
            best_candidate = candidate
    return best_candidate if best_score >= 0.6 else None


def best_value(department: str, label_map: dict[str, int]) -> int | None:
    targets = []
    exact = department_key(department)
    loose = loose_department_key(department)
    for candidate in (exact, loose):
        if candidate and candidate not in targets:
            targets.append(candidate)
    label_variants: dict[str, list[str]] = {}
    for label in label_map:
        variants = [label]
        loose_label = loose_department_key(label)
        if loose_label and loose_label not in variants:
            variants.append(loose_label)
        label_variants[label] = variants

    for target in targets:
        if target in label_map:
            value = label_map[target]
            if plausible_count(value):
                return value
        for label, number in label_map.items():
            if not plausible_count(number):
                continue
            variants = label_variants[label]
            if any(target == variant or target in variant or variant in target for variant in variants):
                return number

    best_number = None
    best_score = 0.0
    for target in targets:
        for label, number in label_map.items():
            if not plausible_count(number):
                continue
            for variant in label_variants[label]:
                score = SequenceMatcher(None, target, variant).ratio()
                if score > best_score:
                    best_score = score
                    best_number = number
    return best_number if best_score >= 0.68 else None


def save_status(**payload: object) -> None:
    STATUS_PATH.parent.mkdir(parents=True, exist_ok=True)
    STATUS_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    target_university = os.environ.get("TARGET_UNIVERSITY", "").strip()
    workbook = openpyxl.load_workbook(WORKBOOK_PATH)
    sheet = workbook[workbook.sheetnames[0]]

    resolved_rows = 0
    save_status(stage="start", source_maps=len(source_maps), resolved_rows=resolved_rows)

    university_rows: dict[str, list[int]] = defaultdict(list)
    for row_index in range(2, sheet.max_row + 1):
        university = sheet.cell(row_index, 3).value
        department = sheet.cell(row_index, 5).value
        if university and department:
            university_rows[str(university)].append(row_index)

    processed_universities = 0
    for university, rows in university_rows.items():
        if target_university and university != target_university:
            continue
        source_maps = build_source_maps(university)
        source_name = best_source(university, source_maps)
        if not source_name:
            continue
        label_map = source_maps[source_name]
        for row_index in rows:
            department = sheet.cell(row_index, 5).value
            value = best_value(str(department), label_map)
            if value is not None:
                sheet.cell(row_index, 9).value = int(value)
                resolved_rows += 1
        processed_universities += 1
        if processed_universities % 10 == 0:
            workbook.save(OUTPUT_PATH)
            save_status(
                stage="progress",
                processed_universities=processed_universities,
                resolved_rows=resolved_rows,
                output=str(OUTPUT_PATH),
            )

    workbook.save(OUTPUT_PATH)
    DONE_PATH.write_text("done", encoding="utf-8")
    try:
        DESKTOP_OUTPUT_PATH.write_bytes(OUTPUT_PATH.read_bytes())
    except Exception:
        pass
    save_status(
        stage="done",
        processed_universities=processed_universities,
        resolved_rows=resolved_rows,
        output=str(OUTPUT_PATH),
    )


if __name__ == "__main__":
    main()
