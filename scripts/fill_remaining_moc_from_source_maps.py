from __future__ import annotations

import importlib.util
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\chris\Desktop\suprema-platform")
INPUT_XLSX = ROOT / "outputs" / "2028_입시계획" / "2028_최종결과.xlsx"
OUTPUT_XLSX = ROOT / "outputs" / "2028_numeric_18_repaired_v8.xlsx"


def load_numeric_module():
    path = ROOT / "scripts" / "numeric_moc_extractor.py"
    spec = importlib.util.spec_from_file_location("numeric_moc_extractor", path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def main() -> None:
    mod = load_numeric_module()
    workbook = load_workbook(INPUT_XLSX)
    sheet = workbook.active

    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 9).value = None

    rows_by_university: dict[str, list[int]] = defaultdict(list)
    for row in range(2, sheet.max_row + 1):
        university = str(sheet.cell(row, 3).value or "").strip()
        if university:
            rows_by_university[university].append(row)

    filled = 0
    debug: list[list[object]] = []

    for university, row_indexes in rows_by_university.items():
        source_maps = mod.build_source_maps(university)
        if not source_maps:
            continue
        source_name = mod.best_source(university, source_maps)
        if not source_name:
            continue
        label_map = source_maps[source_name]
        for row in row_indexes:
            if sheet.cell(row, 9).value not in (None, ""):
                continue
            department = str(sheet.cell(row, 5).value or "").strip()
            value = mod.best_value(department, label_map)
            if value is None:
                continue
            sheet.cell(row, 9).value = int(value)
            filled += 1
            if len(debug) < 30:
                debug.append([row, university, department, value, source_name])

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_XLSX)

    report = {
        "input": str(INPUT_XLSX),
        "output": str(OUTPUT_XLSX),
        "filled": filled,
        "debug": debug,
    }
    (ROOT / "_tmp" / "fill_remaining_moc_from_source_maps.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
