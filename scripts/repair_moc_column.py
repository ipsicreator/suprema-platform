from __future__ import annotations

import json
import re
import sys
import shutil
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

import fitz
import openpyxl

try:
    from rhwp import parse as parse_hwp
except Exception:  # pragma: no cover
    parse_hwp = None


ROOT = Path(r"C:\Users\chris\Desktop\suprema-platform")
INPUT_XLSX = Path(r"C:\Users\chris\Desktop\2028학년도_입시계획_최종_숫자만.xlsx")
OUTPUT_XLSX = ROOT / "outputs" / "2028_numeric_18_repaired.xlsx"
PDF_ROOT = ROOT / "data" / "2028_pdf_extracted"
RAW_ROOT = ROOT / "data" / "pdf_raw_text"

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="ignore")


def normalize(text: object) -> str:
    value = str(text or "")
    value = value.replace("\u0000", "")
    value = re.sub(r"\([^)]*\)", "", value)
    value = re.sub(r"\[[^\]]*\]", "", value)
    value = value.replace("교육대학교", "교대")
    value = value.replace("여자대학교", "여대")
    value = value.replace("과학기술대학교", "과기대")
    value = value.replace("과학기술대", "과기대")
    value = value.replace("대학교", "대")
    value = value.replace("대학", "대")
    value = value.replace("국립", "")
    value = value.replace("캠퍼스", "")
    value = value.replace(" ", "")
    value = value.replace("|", "")
    value = value.replace("ㅣ", "")
    value = value.replace("·", "")
    value = value.replace("-", "")
    return value


def section_kind(text: str) -> str:
    if "정시" in text or "수능" in text:
        return "jungsi"
    if "정원외" in text or "재외국민" in text or "외국인" in text:
        return "outside"
    return "susi"


def numeric_values(text: str) -> list[int]:
    values = []
    for token in re.findall(r"\d{1,3}(?:,\d{3})*|\d+", text):
        try:
            values.append(int(token.replace(",", "")))
        except ValueError:
            continue
    return values


def row_kind(name: str, admission_type: str) -> str:
    text = f"{name or ''} {admission_type or ''}"
    if any(key in text for key in ("정시", "수능", "가군", "나군", "다군", "Ⅰ", "Ⅱ", "Ⅲ")):
        return "jungsi"
    if any(key in text for key in ("농어촌", "특성화고", "교육기회", "기초", "차상위", "한부모", "재외국민", "외국인")):
        return "outside"
    return "susi"


def type_position(name: str, admission_type: str, count_len: int) -> int | None:
    text = f"{name or ''} {admission_type or ''}"
    if count_len <= 0:
        return None

    # Row-specific overrides.
    if "농어촌" in text:
        return 0
    if "교육기회" in text:
        return 1 if count_len > 1 else 0
    if "특성화고" in text:
        return 2 if count_len > 2 else count_len - 1
    if "기초" in text or "차상위" in text or "한부모" in text:
        return 1 if count_len > 1 else 0
    if "신한국인" in text or "한국인" in text:
        return 4 if count_len > 4 else count_len - 1
    if "기회균형" in text or "균형" in text:
        return 3 if count_len > 3 else count_len - 1
    if "사회기여" in text or "기여자" in text:
        return 2 if count_len > 2 else count_len - 1
    if "우수자" in text and "실기" not in text:
        return 1 if count_len > 1 else 0
    if "실기우수자" in text or ("실기" in text and "우수자" in text):
        return count_len - 1
    if "일반전형Ⅲ" in text or "일반전형3" in text or "다군" in text or "Ⅲ" in text:
        return 2 if count_len > 2 else count_len - 1
    if "일반전형Ⅱ" in text or "일반전형2" in text or "나군" in text or "Ⅱ" in text:
        return 1 if count_len > 1 else 0
    if "일반전형Ⅰ" in text or "일반전형1" in text or "가군" in text or "Ⅰ" in text:
        return 0
    if "일반" in text:
        return 0
    if "논술" in text:
        return 0
    return 0


def page_texts_from_json(path: Path) -> list[tuple[int, str]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return [(int(page["page"]), str(page.get("text") or "")) for page in data.get("pages", [])]


def page_texts_from_pdf(path: Path) -> list[tuple[int, str]]:
    doc = fitz.open(path)
    texts = []
    for index, page in enumerate(doc, start=1):
        text = page.get_text("text") or ""
        lines = [text.strip()]
        try:
            finder = page.find_tables()
            for table in finder.tables[:3]:
                try:
                    for row in table.extract():
                        row_text = " | ".join("" if cell is None else str(cell) for cell in row)
                        if row_text.strip():
                            lines.append(row_text)
                except Exception:
                    continue
        except Exception:
            pass
        if len(text.strip()) < 50 and len(lines) == 1:
            # Keep the raw text only for now; OCR is not enabled for this first pass.
            text = page.get_text("blocks") or text
            lines = [text if isinstance(text, str) else str(text)]
        texts.append((index, "\n".join(line for line in lines if line)))
    return texts


def page_texts_from_xlsx(path: Path) -> list[tuple[int, str]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active
    lines = []
    for row in sheet.iter_rows(values_only=True):
        values = [str(value).strip() for value in row if value not in (None, "")]
        if values:
            lines.append(" ".join(values))
    return [(1, "\n".join(lines))]


def page_texts_from_hwp(path: Path) -> list[tuple[int, str]]:
    if parse_hwp is None:
        return []
    try:
        parsed = parse_hwp(str(path))
        text = parsed.extract_text() or ""
        return [(1, text)]
    except Exception:
        return []


def source_pages(path: Path) -> list[tuple[int, str]]:
    if path.suffix.lower() == ".json":
        return page_texts_from_json(path)
    if path.suffix.lower() == ".pdf":
        return page_texts_from_pdf(path)
    if path.suffix.lower() == ".xlsx":
        return page_texts_from_xlsx(path)
    if path.suffix.lower() == ".hwp":
        return page_texts_from_hwp(path)
    return []


def candidate_sources(university: str) -> list[Path]:
    alias_parts = [normalize(university)]
    base = re.sub(r"\([^)]*\)", "", university)
    base_alias = normalize(base)
    if base_alias not in alias_parts:
        alias_parts.append(base_alias)
    if "대학교" in university:
        alias_parts.append(normalize(university.replace("대학교", "대")))
    if "교육대학교" in university:
        alias_parts.append(normalize(university.replace("교육대학교", "교대")))
    if "여자대학교" in university:
        alias_parts.append(normalize(university.replace("여자대학교", "여대")))
    campus_map = {
        "미래": "미래캠퍼스",
        "글로컬": "글로컬",
        "춘천": "춘천캠퍼스",
        "삼척": "삼척캠퍼스",
        "강릉": "강릉캠퍼스",
        "원주": "원주캠퍼스",
        "천안": "천안캠퍼스",
        "세종": "세종캠퍼스",
        "ERICA": "ERICA",
        "WISE": "WISE",
    }
    for key, suffix in campus_map.items():
        if key in university:
            alias_parts.append(normalize(re.sub(r"\([^)]*\)", "", university).replace(key, suffix)))
            alias_parts.append(normalize(re.sub(r"\([^)]*\)", "", university) + suffix))
    files = [p for p in PDF_ROOT.rglob("*") if p.is_file()]
    raw_files = {p.stem: p for p in RAW_ROOT.glob("*.json")}
    scored: list[tuple[float, Path]] = []
    for path in files:
        stem = normalize(path.stem)
        score = 0.0
        for alias in alias_parts:
            if not alias:
                continue
            if stem.startswith(alias):
                score = max(score, 3.0 + len(alias) / 100.0)
            elif alias in stem:
                score = max(score, 2.0 + len(alias) / 100.0)
            else:
                ratio = SequenceMatcher(None, alias, stem).ratio()
                score = max(score, ratio)
        if score > 0:
            scored.append((score, path))
    scored.sort(key=lambda item: (item[0], -len(item[1].name)), reverse=True)

    # Prefer a raw JSON export if we have one for the same university stem.
    for alias in alias_parts:
        for stem, path in raw_files.items():
            if normalize(stem).startswith(alias) or alias.startswith(normalize(stem)):
                scored.insert(0, (10.0, path))
                break

    seen: set[Path] = set()
    result: list[Path] = []
    for _, path in scored:
        if path not in seen:
            seen.add(path)
            result.append(path)
    return result


def section_match_score(page_text: str, target_kind: str) -> int:
    kind = section_kind(page_text)
    if kind == target_kind:
        return 3
    if target_kind == "outside" and kind == "susi":
        return 1
    return 0


def extract_count_for_row(university_pages: list[tuple[int, str]], dept: str, admission_type: str, name: str) -> tuple[int | None, str | None]:
    dept_key = normalize(dept)
    target_kind = row_kind(name, admission_type)
    best: tuple[tuple[int, int, int, int], int | None, str | None] | None = None

    for page_no, text in university_pages:
        lines = [re.sub(r"\s+", " ", line).strip() for line in str(text).splitlines() if line.strip()]
        if not lines:
            continue
        header_text = " ".join(lines[:8])
        if section_match_score(header_text, target_kind) == 0:
            continue
        for line in lines:
            line_key = normalize(line)
            if dept_key not in line_key:
                continue
            nums = numeric_values(line)
            if not nums:
                continue
            position = type_position(name, admission_type, len(nums))
            if position is None:
                continue
            if position >= len(nums):
                position = len(nums) - 1
            value = nums[position]
            # Prefer shorter numeric rows inside the right section.
            score = (
                section_match_score(header_text, target_kind),
                1 if name and normalize(name) in line_key else 0,
                -len(nums),
                -len(line_key),
            )
            candidate = (score, value, line)
            if best is None or candidate[0] > best[0]:
                best = candidate

    if best is None:
        return None, None
    return best[1], best[2]


def main() -> None:
    workbook = openpyxl.load_workbook(INPUT_XLSX)
    sheet = workbook.active

    rows_by_university: dict[str, list[int]] = {}
    for row in range(2, sheet.max_row + 1):
        university = str(sheet.cell(row, 3).value or "").strip()
        if university:
            rows_by_university.setdefault(university, []).append(row)

    source_cache: dict[Path, list[tuple[int, str]]] = {}
    row_updates: dict[int, int] = {}
    debug_rows = []

    for university, row_indexes in rows_by_university.items():
        sources = candidate_sources(university)
        if not sources:
            continue

        pages: list[tuple[int, str]] = []
        for source in sources[:3]:
            if source not in source_cache:
                try:
                    source_cache[source] = source_pages(source)
                except Exception:
                    source_cache[source] = []
            pages = source_cache[source]
            if pages:
                break
        if not pages:
            continue

        for row in row_indexes:
            dept = str(sheet.cell(row, 5).value or "").strip()
            admission_type = str(sheet.cell(row, 6).value or "").strip()
            name = str(sheet.cell(row, 7).value or "").strip()
            count, matched_line = extract_count_for_row(pages, dept, admission_type, name)
            if count is not None:
                row_updates[row] = count
                if len(debug_rows) < 20:
                    debug_rows.append((row, university, dept, admission_type, name, count, matched_line))

    for row, value in row_updates.items():
        sheet.cell(row, 9).value = value

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_XLSX)

    report = {
        "input": str(INPUT_XLSX),
        "output": str(OUTPUT_XLSX),
        "updated_rows": len(row_updates),
        "debug": debug_rows,
    }
    (ROOT / "_tmp").mkdir(exist_ok=True)
    (ROOT / "_tmp" / "repair_moc_column_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
