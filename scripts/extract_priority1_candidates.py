from __future__ import annotations

import json
import re
import subprocess
from pathlib import Path

import fitz
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font


ROOT = Path(r"C:\Users\chris\Desktop\suprema-platform")
SUMMARY_JSON = ROOT / "outputs" / "adiga_2028_keyword_pages_priority1.json"
OUTPUT_XLSX = ROOT / "outputs" / "adiga_2028_priority1_candidates.xlsx"
TMP_DIR = ROOT / ".tmp" / "adiga_ocr_pages_extract"
TESSERACT = Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe")

KEYWORDS = ("모집인원", "모집단위", "입학정원", "전형명")


def ocr_page(page: fitz.Page, pdf_stem: str, page_index: int) -> str:
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    image_path = TMP_DIR / f"{pdf_stem}_p{page_index + 1}.png"
    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
    pix.save(image_path)
    try:
        result = subprocess.run(
            [str(TESSERACT), str(image_path), "stdout", "-l", "kor+eng"],
            capture_output=True,
            text=False,
            check=False,
        )
        return (result.stdout or b"").decode("utf-8", errors="ignore")
    finally:
        if image_path.exists():
            image_path.unlink()


def normalize(text: str) -> str:
    text = text or ""
    text = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def safe_cell(value):
    if isinstance(value, str):
        value = ILLEGAL_CHARACTERS_RE.sub(" ", value)
        value = "".join(
            ch
            for ch in value
            if ch in "\t\n\r"
            or 0x20 <= ord(ch) <= 0xD7FF
            or 0xE000 <= ord(ch) <= 0xFFFD
            or 0x10000 <= ord(ch) <= 0x10FFFF
        )
        return value
    return value


def extract_candidates(text: str) -> list[str]:
    text = normalize(text)
    candidates: list[str] = []
    patterns = [
        r"모집인원[^0-9]{0,20}([0-9][0-9,]*)",
        r"모집인원[^0-9]{0,40}([0-9][0-9,]*)",
        r"모집인원\s*[:：]?\s*([0-9][0-9,]*)",
        r"입학정원[^0-9]{0,20}([0-9][0-9,]*)",
        r"모집단위[^0-9]{0,20}([0-9][0-9,]*)",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, text):
            value = match.group(1).replace(",", "")
            if value.isdigit():
                number = int(value)
                if 1900 <= number <= 2100:
                    continue
                candidates.append(str(number))
        if candidates:
            break
    # de-duplicate while preserving order
    seen = set()
    unique: list[str] = []
    for value in candidates:
        if value not in seen:
            seen.add(value)
            unique.append(value)
    return unique


def choose_value(candidates: list[str]) -> str:
    if len(candidates) == 1:
        return candidates[0]
    return ""


def main() -> None:
    summary = json.loads(SUMMARY_JSON.read_text(encoding="utf-8"))
    workbook = Workbook()
    ws_summary = workbook.active
    ws_summary.title = "summary"
    ws_pages = workbook.create_sheet("pages")
    ws_review = workbook.create_sheet("review")

    summary_headers = [
        "school",
        "source_files",
        "bundle_pages",
        "bundle_path",
        "hit_pages_total",
        "pages_with_moc",
        "first_value",
        "review_pages",
    ]
    page_headers = [
        "school",
        "member",
        "page",
        "keywords",
        "candidates",
        "selected",
        "source_kind",
        "snippet",
    ]
    review_headers = [
        "school",
        "member",
        "page",
        "issue",
        "candidates",
        "snippet",
    ]

    ws_summary.append(summary_headers)
    ws_pages.append(page_headers)
    ws_review.append(review_headers)

    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
    for cell in ws_pages[1]:
        cell.font = Font(bold=True)
    for cell in ws_review[1]:
        cell.font = Font(bold=True)

    summary_rows = []
    page_rows = []
    review_rows = []

    for school_entry in summary:
        school = school_entry["school"]
        bundle_value = school_entry.get("bundle_path") or ""
        if not bundle_value:
            summary_rows.append(
                [
                    school,
                    school_entry.get("source_files", 0),
                    school_entry.get("bundle_pages", 0),
                    "",
                    0,
                    0,
                    "",
                    "no_bundle",
                ]
            )
            continue
        bundle_path = Path(bundle_value)
        if not bundle_path.exists():
            summary_rows.append(
                [
                    school,
                    school_entry.get("source_files", 0),
                    school_entry.get("bundle_pages", 0),
                    str(bundle_path),
                    0,
                    0,
                    "",
                    "no_bundle",
                ]
            )
            continue

        bundle = fitz.open(bundle_path)
        hit_pages_total = 0
        pages_with_moc = 0
        first_value = ""
        review_pages = 0

        for page_index in range(bundle.page_count):
            page = bundle[page_index]
            text = page.get_text("text")
            source_kind = "text"
            if not text.strip():
                text = ocr_page(page, bundle_path.stem, page_index)
                source_kind = "ocr"

            text = normalize(text)
            keywords = [keyword for keyword in KEYWORDS if keyword in text]
            if not keywords:
                continue

            hit_pages_total += 1
            snippet_index = min((text.find(keyword) for keyword in keywords if text.find(keyword) >= 0), default=0)
            snippet = text[max(0, snippet_index - 160): snippet_index + 240]
            candidates = extract_candidates(text if "모집인원" in keywords else snippet)
            selected = choose_value(candidates)
            if "모집인원" in keywords:
                pages_with_moc += 1
                if not first_value and selected:
                    first_value = selected
                if len(candidates) != 1:
                    review_pages += 1
                    review_rows.append(
                        [
                            school,
                            school_entry["files"][0]["member"] if school_entry.get("files") else "",
                            page_index + 1,
                            "ambiguous" if candidates else "no_candidate",
                            ", ".join(candidates),
                            snippet,
                        ]
                    )

            page_rows.append(
                [
                    school,
                    school_entry["files"][0]["member"] if school_entry.get("files") else "",
                    page_index + 1,
                    ", ".join(keywords),
                    ", ".join(candidates),
                    selected,
                    source_kind,
                    snippet,
                ]
            )

        summary_rows.append(
            [
                school,
                school_entry.get("source_files", 0),
                school_entry.get("bundle_pages", 0),
                str(bundle_path),
                hit_pages_total,
                pages_with_moc,
                first_value,
                review_pages,
            ]
        )

    for row in summary_rows:
        ws_summary.append([safe_cell(value) for value in row])
    for row in page_rows:
        ws_pages.append([safe_cell(value) for value in row])
    for row in review_rows:
        ws_review.append([safe_cell(value) for value in row])

    for ws in [ws_summary, ws_pages, ws_review]:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # basic column widths
    widths = {
        "summary": [24, 12, 12, 72, 14, 14, 12, 14],
        "pages": [20, 36, 8, 18, 18, 12, 10, 100],
        "review": [20, 36, 8, 16, 18, 100],
    }
    for ws, width_list in [(ws_summary, widths["summary"]), (ws_pages, widths["pages"]), (ws_review, widths["review"])]:
        for idx, width in enumerate(width_list, start=1):
            ws.column_dimensions[chr(64 + idx)].width = width

    workbook.save(OUTPUT_XLSX)
    print(json.dumps(
        {
            "output": str(OUTPUT_XLSX),
            "summary_rows": len(summary_rows),
            "page_rows": len(page_rows),
            "review_rows": len(review_rows),
        },
        ensure_ascii=False,
        indent=2,
    ))


if __name__ == "__main__":
    main()
